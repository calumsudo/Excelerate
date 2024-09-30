import io
import os
import openpyxl
from datetime import datetime, timedelta
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from log import log_to_file


def get_workbook_data(workbook_bytes, selected_file, output_path, portfolio_name):
    if isinstance(workbook_bytes, list):
        workbook_bytes = b"".join(workbook_bytes)

    portfolio_name = {1: "Alder", 2: "White Rabbit"}.get(portfolio_name, portfolio_name)

    workbook = load_workbook(filename=io.BytesIO(workbook_bytes))

    current_date = datetime.now()
    date_string = current_date.strftime("%m_%d_%Y")
    os.makedirs(
        os.path.expanduser(f"{output_path}/{portfolio_name}/{date_string}"),
        exist_ok=True,
    )
    file_name = selected_file.removesuffix(".xlsx")
    file_path_backup = os.path.expanduser(
        f"{output_path}/{portfolio_name}/{date_string}/{file_name}_BACKUP_{date_string}.xlsx"
    )

    with open(file_path_backup, "wb") as file:
        file.write(workbook_bytes)

    return workbook

def add_net_rtr_column_if_needed(worksheet, selected_date=None, header_row=2):
    # Format the date as 'M/D'
    month_day = (
        selected_date.strftime("%-m/%-d")
        if selected_date
        else (datetime.now() + timedelta((4 - datetime.now().weekday()) % 7)).strftime(
            "%-m/%-d"
        )
    )

    # Find the column index for 'R&H Net RTR Balance'
    rtr_balance_column = None
    for idx, cell in enumerate(worksheet[header_row]):
        if cell.value and "R&H Net RTR Balance" in cell.value:
            rtr_balance_column = idx + 1  # Excel columns are 1-indexed
            break

    if rtr_balance_column is None:
        raise ValueError("R&H Net RTR Balance column not found in the sheet.")

    # Check if the Net RTR column for the current date already exists
    net_rtr_column_index = None
    for idx, cell in enumerate(worksheet[header_row]):
        if cell.value and f"Net RTR {month_day}" == cell.value:
            net_rtr_column_index = idx + 1  # Excel columns are 1-indexed
            break

    # If it doesn't exist, insert a new column
    if net_rtr_column_index is None:
        worksheet.insert_cols(rtr_balance_column)
        net_rtr_column_index = rtr_balance_column

        worksheet.cell(row=header_row, column=net_rtr_column_index).value = (
            f"Net RTR {month_day}"
        )
        worksheet.cell(row=1, column=net_rtr_column_index).value = worksheet.title

        # Copy formatting from the adjacent cell
        for row in range(1, worksheet.max_row + 1):
            left_cell = worksheet.cell(row=row, column=net_rtr_column_index - 1)
            new_cell = worksheet.cell(row=row, column=net_rtr_column_index)

            new_cell.font = copy(left_cell.font)
            new_cell.border = copy(left_cell.border)
            new_cell.fill = copy(left_cell.fill)
            new_cell.number_format = copy(left_cell.number_format)
            new_cell.protection = copy(left_cell.protection)
            new_cell.alignment = copy(left_cell.alignment)

    added_col = get_column_letter(net_rtr_column_index)
    update_total_net_rtr_formula(worksheet, added_col, header_row)

    return added_col


def update_total_net_rtr_formula(worksheet, net_rtr_column, header_row=2):
    total_net_rtr_header = "Total Net RTR Payment Received"
    total_net_rtr_column = None
    for cell in worksheet[header_row]:
        if cell.value == total_net_rtr_header:
            total_net_rtr_column = get_column_letter(cell.column)
            break

    if total_net_rtr_column is None:
        return

    start_column = get_column_letter(openpyxl.utils.column_index_from_string(total_net_rtr_column) + 1)
    for row in range(header_row + 1, worksheet.max_row + 1):
        formula = f"=SUM({start_column}{row}:{net_rtr_column}{row})"
        worksheet[f"{total_net_rtr_column}{row}"].value = formula


def map_net_amount_to_excel(
    worksheet, df_csv, net_rtr_column, output_path, portfolio_name, header_row=2
):
    # Build a mapping from Funder Advance ID to row number in the worksheet
    advance_id_to_row = {}
    for row_cells in worksheet.iter_rows(min_row=header_row + 1, min_col=5, max_col=5):
        cell = row_cells[0]  # Column 5 corresponds to 'Funder Advance ID'
        if cell.value:
            advance_id = str(cell.value).strip()
            advance_id_to_row[advance_id] = cell.row

    for _, row in df_csv.iterrows():
        advance_id = str(row["Funder Advance ID"]).strip()
        if advance_id == "All":
            continue

        net_amount = row["Sum of Syn Net Amount"]
        if advance_id in advance_id_to_row:
            excel_row_num = advance_id_to_row[advance_id]
            worksheet[f"{net_rtr_column}{excel_row_num}"].value = net_amount
            log_to_file(
                f"Advance ID '{advance_id}' mapped to row {excel_row_num} with net amount {net_amount}.",
                output_path,
                portfolio_name,
            )
        else:
            log_to_file(
                f"Funder Advance ID '{advance_id}' not found in Excel sheet.",
                output_path,
                portfolio_name,
            )


def add_data_to_sheet(
    workbook, df, sheet_name, output_path, portfolio_name, selected_date=None
):
    try:
        sheet = workbook[sheet_name]
        log_to_file(f"Adding data to sheet '{sheet_name}'...", output_path, portfolio_name)
        print(f"Adding data to sheet '{sheet_name}'...")  # Debugging statement

        # Extract Advance IDs and Merchant Names from DataFrame
        df_advance_ids = (
            df[["Funder Advance ID", "Merchant Name"]]
            .dropna(subset=["Funder Advance ID"])
            .set_index("Funder Advance ID")
            .to_dict("index")
        )
        print(f"df_advance_ids: {df_advance_ids}")  # Debugging statement

                # Create a mapping of 'Funder Advance ID' to row number in the worksheet
        worksheet_advance_ids = {}
        header_row = 2  # Adjust if your header is on a different row
        for row_cells in sheet.iter_rows(min_row=header_row + 1, min_col=5, max_col=5):
            cell = row_cells[0]  # Column 5 corresponds to 'Funder Advance ID'
            if cell.value:
                advance_id = str(cell.value).strip()
                worksheet_advance_ids[advance_id] = cell.row
        print(f"worksheet_advance_ids: {worksheet_advance_ids}")

        # Find unmatched Advance IDs and corresponding Merchant Names
        unmatched_advance_ids = set(df_advance_ids) - set(worksheet_advance_ids)
        unmatched_info = {
            aid: df_advance_ids[aid]["Merchant Name"]
            for aid in unmatched_advance_ids
            if aid != "All" and df_advance_ids[aid]["Merchant Name"] != "All"
        }

        # Format unmatched info for logging
        formatted_unmatched_info = ", ".join(
            f"{aid}: {info}" for aid, info in unmatched_info.items()
        )
        detailed_unmatched_info = [
            {"sheet_name": sheet_name, "advance_id": aid, "merchant_name": info}
            for aid, info in unmatched_info.items()
        ]

        log_to_file(
            f"Advance IDs and Merchants not found in Excel sheet: {formatted_unmatched_info}",
            output_path,
            portfolio_name,
        )

        # Ensure the Net RTR column is added and get its column letter
        net_rtr_column = add_net_rtr_column_if_needed(sheet, selected_date)
        print(f"net_rtr_column: {net_rtr_column}")  # Debugging statement

        # Updated to map using 'Advance ID'
        map_net_amount_to_excel(sheet, df, net_rtr_column, output_path, portfolio_name)

        # Save changes to the workbook
        output_bytes = BytesIO()
        workbook.save(output_bytes)
        output_bytes.seek(0)
        final_bytes = output_bytes.getvalue()
        print(f"final_bytes length: {len(final_bytes)}")  # Debugging statement
        if not final_bytes:
            raise ValueError("final_bytes is None or empty")

        return final_bytes, detailed_unmatched_info
    except Exception as e:
        error_message = f"Error in add_data_to_sheet: {str(e)}"
        print(error_message)
        log_to_file(error_message, output_path, portfolio_name)
        return None, []