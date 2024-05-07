import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Function to add a column if the current Friday's Net RTR column does not exist
def add_net_rtr_column_if_needed(worksheet, header_row=2):
    current_friday = (datetime.now() + timedelta((4 - datetime.now().weekday()) % 7))
    current_friday = f"{current_friday.month}/{current_friday.day}"
    rtr_balance_column = None
    net_rtr_column_index = None
    
    # Find the R&H Net RTR Balance column index
    for idx, cell in enumerate(worksheet[header_row]):
        if cell.value and "R&H Net RTR Balance" in cell.value:
            rtr_balance_column = idx + 1  # Excel is 1-indexed
            break

    # Check if the column for the current Friday already exists
    for idx, cell in enumerate(worksheet[header_row]):
        if cell.value and f'Net RTR {current_friday}' == cell.value:
            net_rtr_column_index = idx + 1  # Excel is 1-indexed
            break

    # If no RTR balance column found, something is wrong with the sheet format
    if rtr_balance_column is None:
        raise ValueError("R&H Net RTR Balance column not found in the sheet.")

    # If the Net RTR column for the current Friday doesn't exist, we insert it before the R&H Net RTR Balance column
    if net_rtr_column_index is None:
        worksheet.insert_cols(rtr_balance_column)
        net_rtr_column_index = rtr_balance_column
        # Set the value for the header of the new column
        worksheet.cell(row=header_row, column=rtr_balance_column).value = f'Net RTR {current_friday}'
        
    return get_column_letter(net_rtr_column_index)

def update_total_net_rtr_formula(worksheet, net_rtr_column):
    total_net_rtr_header = "Total Net RTR Payment Received"
    header_row = 2  # Assuming headers are in the second row
    
    # Iterate through cells in the header row without using values_only=True
    for cell in worksheet[header_row]:
        # Now cell is a Cell object, and you can access cell.value
        if cell.value == total_net_rtr_header:
            # Now correctly access the cell's attributes
            # Find the column letter for the cell
            col_letter = openpyxl.utils.get_column_letter(cell.column)
            # Update the formula for the entire column starting from the next row of the header
            for row in range(header_row + 1, worksheet.max_row + 1):
                formula_cell = f'{col_letter}{row}'
                # Assuming you want to sum from AG (column 33) to the new Net RTR column for each row
                start_column_letter = openpyxl.utils.get_column_letter(33)  # AG column
                end_column_letter = net_rtr_column
                worksheet[formula_cell].value = f'=SUM({start_column_letter}{row}:{end_column_letter}{row})'
            break  # Exit loop after updating

def map_net_amount_to_excel(worksheet, df_csv, net_rtr_column, header_row=2):
    advance_id_to_row = {}
    # Adjust to column 'E' for 'Funder Advance ID'
    for row in worksheet.iter_rows(min_row=header_row + 1, min_col=5, max_col=5):
        cell = row[0]
        if cell.value:
            # Convert cell value to string, strip whitespace, and store it with its row number
            advance_id_to_row[str(cell.value).strip()] = cell.row

    # Mapping using 'Funder Advance ID'
    for index, row in df_csv.iterrows():
        advance_id = str(row['Funder Advance ID']).strip()

        # Skip processing if 'Funder Advance ID' is 'All'
        if advance_id == 'All':
            print(f"Skipping Funder Advance ID '{advance_id}' as it is designated to ignore.")
            continue

        net_amount = row['Sum of Syn Net Amount']
        
        if advance_id in advance_id_to_row:
            excel_row_num = advance_id_to_row[advance_id]
            cell_reference = f'{net_rtr_column}{excel_row_num}'
            worksheet[cell_reference].value = net_amount
            print("Mapped:", advance_id, "to row:", excel_row_num, "with net amount:", net_amount)
        else:
            print(f"Funder Advance ID '{advance_id}' not found in Excel sheet.")


            
# Function to process CSV and Excel data
def process_csv_and_excel(csv_file_path, excel_file_path):
    # Load the CSV into a DataFrame
    df_csv = pd.read_csv(csv_file_path)

    # Open the Excel workbook
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb['Kings']

    # Add Net RTR column if needed and get the letter of the Net RTR column
    net_rtr_column = add_net_rtr_column_if_needed(sheet)

    # Map 'Sum of Syn Net Amount' from CSV to Excel
    map_net_amount_to_excel(sheet, df_csv, net_rtr_column)

    # Save the Excel file
    wb.save(excel_file_path)

    return df_csv


# Function to create log file for unmatched merchants
def create_log_for_unmatched(unmatched_merchants, log_file_path):
    with open(log_file_path, 'w') as file:
        for merchant, amount in unmatched_merchants.items():
            file.write(f'{merchant}: {amount}\n')

# Inside your main function call
if __name__ == "__main__":
    # Assuming 'Ass.xlsx' is the correct Excel file name and it's in the same directory as your script
    excel_file_path = 'ALDER2.xlsx'
    csv_file_path = 'KINGS2.csv'

    unmatched = process_csv_and_excel(csv_file_path, excel_file_path)
    
    # Load the workbook and sheet again
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb['Kings']
    
    # Get the column letter for the new Net RTR column
    net_rtr_column = add_net_rtr_column_if_needed(sheet)
    df_csv = process_csv_and_excel(csv_file_path, excel_file_path)

    # Map the amounts from the CSV to the Excel file
    map_net_amount_to_excel(sheet, df_csv, net_rtr_column)

    # Now call the function to update the formula
    update_total_net_rtr_formula(sheet, net_rtr_column)

    # Don't forget to save the workbook after making changes
    wb.save(excel_file_path)

    if not unmatched.empty:
        create_log_for_unmatched(unmatched, 'loggy.txt')


