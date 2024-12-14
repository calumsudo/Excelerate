# app/core/data_processing/excel/workbook_manager.py

import openpyxl
from pathlib import Path
from datetime import datetime
from copy import copy
from typing import Dict, List, Tuple, Optional
import pandas as pd
import logging
from openpyxl.utils import get_column_letter

class WorkbookManager:
    # Mapping between parser names and worksheet names
    SHEET_MAPPING = {
        "ACS": "ACS",
        "BHB": "BHB",
        "Boom": "Boom",
        "ClearView": "CV",
        "EFIN": "EFin",
        "Kings": "Kings",
        "Vesper": "VSPR"
    }

    def __init__(self, file_manager):
        self.file_manager = file_manager
        self.logger = logging.getLogger(__name__)

    def backup_workbook(self, portfolio_path: Path, friday_date: datetime) -> Path:
        """Create a backup of the workbook before modifications"""
        try:
            source_path = portfolio_path
            backup_name = f"{portfolio_path.stem}_backup_{friday_date.strftime('%Y%m%d')}.xlsx"
            backup_path = portfolio_path.parent / backup_name
            
            import shutil
            shutil.copy2(source_path, backup_path)
            
            self.logger.info(f"Created workbook backup at {backup_path}")
            return backup_path
            
        except Exception as e:
            self.logger.error(f"Failed to create workbook backup: {str(e)}")
            raise

    def _add_net_rtr_column(self, worksheet, friday_date: datetime, header_row: int = 2) -> str:
        """Add Net RTR column for the current date if it doesn't exist"""
        month_day = friday_date.strftime("%-m/%-d")
        
        # Find RTR Balance column
        rtr_balance_col = None
        for idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value and "R&H Net RTR Balance" in str(cell.value):
                rtr_balance_col = idx
                break
                
        if not rtr_balance_col:
            raise ValueError("R&H Net RTR Balance column not found")
            
        # Check if column already exists
        net_rtr_col = None
        for idx, cell in enumerate(worksheet[header_row], 1):
            if cell.value and f"Net RTR {month_day}" == str(cell.value):
                net_rtr_col = idx
                break
                
        # Add new column if needed
        if not net_rtr_col:
            worksheet.insert_cols(rtr_balance_col)
            net_rtr_col = rtr_balance_col
            
            # Set header values
            worksheet.cell(row=header_row, column=net_rtr_col).value = f"Net RTR {month_day}"
            worksheet.cell(row=1, column=net_rtr_col).value = worksheet.title
            
            # Copy formatting from adjacent column
            for row in range(1, worksheet.max_row + 1):
                source = worksheet.cell(row=row, column=net_rtr_col - 1)
                target = worksheet.cell(row=row, column=net_rtr_col)
                
                target.font = copy(source.font)
                target.border = copy(source.border)
                target.fill = copy(source.fill)
                target.number_format = copy(source.number_format)
                target.protection = copy(source.protection)
                target.alignment = copy(source.alignment)
                
        return get_column_letter(net_rtr_col)

    def _update_total_formula(self, worksheet, net_rtr_col: str, header_row: int = 2):
        """Update the Total Net RTR Payment Received formula"""
        total_col = None
        for cell in worksheet[header_row]:
            if cell.value and "Total Net RTR Payment Received" in str(cell.value):
                total_col = get_column_letter(cell.column)
                break
                
        if total_col:
            start_col = get_column_letter(openpyxl.utils.column_index_from_string(total_col) + 1)
            for row in range(header_row + 1, worksheet.max_row + 1):
                worksheet[f"{total_col}{row}"].value = f"=SUM({start_col}{row}:{net_rtr_col}{row})"

    def update_workbook(self, 
                       portfolio_path: Path,
                       pivot_data: pd.DataFrame,
                       funder: str,
                       friday_date: datetime) -> Tuple[List[Dict], Optional[str]]:
        """
        Update workbook with new net values from pivot table.
        
        Returns:
            Tuple containing:
            - List of unmatched advance IDs with merchant names
            - Error message if any
        """
        try:
            # Get sheet name from mapping
            sheet_name = self.SHEET_MAPPING.get(funder)
            if not sheet_name:
                raise ValueError(f"No sheet mapping found for funder {funder}")

            # Load workbook
            workbook = openpyxl.load_workbook(portfolio_path)
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet {sheet_name} not found in workbook")

            worksheet = workbook[sheet_name]
            
            # Add/get Net RTR column
            net_rtr_col = self._add_net_rtr_column(worksheet, friday_date)
            
            # Update total formula
            self._update_total_formula(worksheet, net_rtr_col)
            
            # Build mapping of Advance IDs to rows
            advance_id_map = {}
            header_row = 2  # Assuming header is always row 2
            for row in worksheet.iter_rows(min_row=header_row + 1, min_col=5, max_col=5):
                cell = row[0]
                if cell.value:
                    advance_id_map[str(cell.value).strip()] = cell.row
                    
            # Track unmatched IDs
            unmatched = []
            
            # Update net values
            for _, row in pivot_data.iterrows():
                advance_id = str(row["Advance ID"]).strip()
                if advance_id == "Totals":  # Skip totals row
                    continue

                net_value = row["Sum of Syn Net Amount"]
                if net_value == 0:
                    continue
                    
                if advance_id in advance_id_map:
                    excel_row = advance_id_map[advance_id]
                    worksheet[f"{net_rtr_col}{excel_row}"].value = net_value
                else:
                    unmatched.append({
                        "sheet_name": sheet_name,
                        "advance_id": advance_id,
                        "merchant_name": row["Merchant Name"]
                    })
                    
            # Save workbook
            workbook.save(portfolio_path)
            
            self.logger.info(f"Updated {sheet_name} worksheet with {len(pivot_data) - len(unmatched)} matches "
                           f"and {len(unmatched)} unmatched IDs")
            
            return unmatched, None
            
        except Exception as e:
            error_msg = f"Error updating workbook: {str(e)}"
            self.logger.error(error_msg)
            return [], error_msg