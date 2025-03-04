# app/core/data_processing/parsers/big_parser.py

from pathlib import Path
import pandas as pd
from typing import Tuple, Optional
from .base_parser import BaseParser
import openpyxl
import logging
import warnings
import re

# Suppress the specific openpyxl warning about default styles
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message="Workbook contains no default style, apply openpyxl's default",
)
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message="Data Validation extension is not supported and will be removed",
)


class BIGParser(BaseParser):
    def __init__(self, file_path: Path):
        super().__init__(file_path)
        self.funder_name = "BIG"
        # We won't validate columns exactly with required_columns since
        # the BIG file format has many columns and we only need a few
        self.required_columns = []
        self.column_types = {}

        # Initialize logger
        self.logger = logging.getLogger(f"parser.{self.__class__.__name__}")

    def validate_format(self) -> Tuple[bool, str]:
        """Validate BIG report format by checking for required worksheets and columns"""
        try:
            # Ensure file is Excel format
            if self.file_path.suffix.lower() not in [".xlsx", ".xls"]:
                return False, "File must be an Excel file (.xlsx or .xls)"

            self.logger.info(f"Loading workbook: {self.file_path}")

            # Load workbook
            workbook = openpyxl.load_workbook(self.file_path, read_only=True)

            # Log all available sheets for debugging
            self.logger.info(f"Available sheets in workbook: {workbook.sheetnames}")

            # More flexible sheet name matching
            alder_match = next((s for s in workbook.sheetnames if "R&H" in s), None)
            wr_match = next(
                (s for s in workbook.sheetnames if "White Rabbit" in s), None
            )

            if alder_match:
                self.logger.info(f"Found Alder sheet: {alder_match}")
            if wr_match:
                self.logger.info(f"Found White Rabbit sheet: {wr_match}")

            if not alder_match and not wr_match:
                return (
                    False,
                    f"Could not find any portfolio sheets. Available sheets: {workbook.sheetnames}",
                )

            return True, ""

        except Exception as e:
            return False, f"Error validating format: {str(e)}"

    def detect_portfolio(self) -> Optional[str]:
        """Detect which portfolio this BIG file is for based on available sheets"""
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True)

            # More flexible sheet name matching
            for sheet in workbook.sheetnames:
                if "R&H" in sheet:
                    self.logger.info(f"Detected Alder portfolio from sheet: {sheet}")
                    self.alder_sheet_name = sheet
                    return "Alder"
                elif "White Rabbit" in sheet:
                    self.logger.info(
                        f"Detected White Rabbit portfolio from sheet: {sheet}"
                    )
                    self.wr_sheet_name = sheet
                    return "White Rabbit"

            self.logger.warning(
                f"Could not detect portfolio from sheets: {workbook.sheetnames}"
            )
            return None

        except Exception as e:
            self.logger.error(f"Error detecting portfolio: {str(e)}")
            return None

    def get_portfolio_sheet_name(self, portfolio: str) -> Optional[str]:
        """Get the correct sheet name for the given portfolio"""
        try:
            workbook = openpyxl.load_workbook(self.file_path, read_only=True)

            # Use stored sheet names if available
            if portfolio == "Alder" and hasattr(self, "alder_sheet_name"):
                return self.alder_sheet_name
            elif portfolio == "White Rabbit" and hasattr(self, "wr_sheet_name"):
                return self.wr_sheet_name

            # Otherwise search for matching sheets
            for sheet in workbook.sheetnames:
                if portfolio == "Alder" and "R&H" in sheet:
                    self.alder_sheet_name = sheet
                    return sheet
                elif portfolio == "White Rabbit" and "White Rabbit" in sheet:
                    self.wr_sheet_name = sheet
                    return sheet

            self.logger.warning(f"No sheet found for portfolio: {portfolio}")
            return None

        except Exception as e:
            self.logger.error(f"Error getting portfolio sheet name: {str(e)}")
            return None

    def clean_advance_id(self, value):
        """Clean and format Advance ID to ensure it's a proper string without decimals"""
        try:
            if pd.isna(value) or value is None:
                return None

            # Convert to string and strip whitespace
            id_str = str(value).strip()

            # Check if empty after stripping
            if not id_str:
                return None

            # If it's a float with decimals (like 1234.0), convert to integer string
            try:
                if "." in id_str:
                    # Try to convert to float first to handle scientific notation
                    float_val = float(id_str)
                    # Check if it's actually an integer (like 1234.0)
                    if float_val.is_integer():
                        return str(int(float_val))
                    # Otherwise keep the decimal version as string
                    return id_str
                # Try to convert to int if it looks like a number
                elif id_str.isdigit():
                    return id_str
                else:
                    return id_str
            except ValueError:
                # If conversion fails, just return the original string
                return id_str

        except Exception as e:
            self.logger.error(f"Error cleaning Advance ID: {str(e)}")
            return str(value)

    def parse_sum_formula(self, formula):
        """Parse a SUM formula and extract the cell range"""
        if not formula or not isinstance(formula, str):
            return None

        # Check if it's a SUM formula
        sum_match = re.match(r"=SUM\(([A-Z]+\d+):([A-Z]+\d+)\)", formula)
        if sum_match:
            start_cell = sum_match.group(1)
            end_cell = sum_match.group(2)
            self.logger.info(f"Found SUM formula: {start_cell} to {end_cell}")
            return (start_cell, end_cell)

        return None

    def process_with_formula_handling(self, portfolio: str) -> pd.DataFrame:
        """Process BIG data with special handling for formulas"""
        try:
            # Get the correct sheet name for this portfolio
            sheet_name = self.get_portfolio_sheet_name(portfolio)
            if not sheet_name:
                raise ValueError(f"Unable to find sheet for portfolio: {portfolio}")

            # Load workbook with data_only=True to evaluate formulas
            self.logger.info(f"Loading workbook with data_only=True: {self.file_path}")
            workbook_calculated = openpyxl.load_workbook(self.file_path, data_only=True)

            # Also load a version without formula evaluation to examine formulas
            workbook_formulas = openpyxl.load_workbook(self.file_path)

            # Get both worksheet versions
            worksheet_calculated = workbook_calculated[sheet_name]
            worksheet_formulas = workbook_formulas[sheet_name]

            # Lists to store data
            advance_ids = []
            merchant_names = []
            net_amounts = []

            # Get columns A and C first - these don't involve formulas
            id_col_letter = "A"  # Column A for Funding ID
            name_col_letter = "C"  # Column C for Merchant Name

            # Find header row first
            start_row = 3  # Default start at row 3
            header_values = [
                "funding id",
                "fundingid",
                "funding_id",
                "id",
                "advance id",
                "advanceid",
            ]

            for row_idx in range(1, 10):  # Check first 10 rows
                id_cell = worksheet_calculated[f"{id_col_letter}{row_idx}"]
                if id_cell.value and str(id_cell.value).lower() in header_values:
                    start_row = row_idx + 1
                    self.logger.info(
                        f"Found header row at {row_idx}, data starts at row {start_row}"
                    )
                    break

            # First check column AI for formula patterns
            self.logger.info("Examining formulas in column AI")

            # Check a few rows in column AI to see if they have SUM formulas
            formula_patterns = []
            for row_idx in range(start_row, start_row + 5):
                cell = worksheet_formulas[f"AI{row_idx}"]
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and cell.value.startswith("=SUM(")
                ):
                    formula_range = self.parse_sum_formula(cell.value)
                    if formula_range:
                        formula_patterns.append(formula_range)

            # Determine source columns based on formulas
            source_cols = []
            if formula_patterns:
                # Extract column letters from the first formula pattern
                self.logger.info(f"Found formula patterns: {formula_patterns}")
                if formula_patterns[0]:
                    start_cell, end_cell = formula_patterns[0]
                    start_col = re.match(r"([A-Z]+)\d+", start_cell).group(1)
                    end_col = re.match(r"([A-Z]+)\d+", end_cell).group(1)

                    # Generate all columns in the range
                    current_col = start_col
                    while True:
                        source_cols.append(current_col)
                        if current_col == end_col:
                            break
                        # Increment column letter
                        current_col = openpyxl.utils.get_column_letter(
                            openpyxl.utils.column_index_from_string(current_col) + 1
                        )

                    self.logger.info(
                        f"Will use source columns for calculation: {source_cols}"
                    )

            # Process rows
            row_count = 0
            data_rows = 0

            for row_idx in range(start_row, worksheet_calculated.max_row + 1):
                row_count += 1

                # Get cells for this row
                id_cell = worksheet_calculated[f"{id_col_letter}{row_idx}"]
                name_cell = worksheet_calculated[f"{name_col_letter}{row_idx}"]

                # Skip rows with empty ID
                if id_cell.value is None or str(id_cell.value).strip() == "":
                    continue

                # Skip header rows
                if id_cell.value and str(id_cell.value).lower() in header_values:
                    continue

                # Extract and clean values
                advance_id = self.clean_advance_id(id_cell.value)
                if advance_id is None:
                    continue

                merchant_name = str(name_cell.value).strip() if name_cell.value else ""

                # First try to get evaluated formula value from column AI
                amount_cell = worksheet_calculated[f"AI{row_idx}"]
                amount = amount_cell.value if amount_cell.value is not None else 0.0

                # If amount is 0 or None and we have source columns, calculate ourselves
                if (amount is None or amount == 0) and source_cols:
                    amount = 0.0
                    for col in source_cols:
                        cell = worksheet_calculated[f"{col}{row_idx}"]
                        if cell.value is not None:
                            try:
                                amount += float(cell.value)
                            except (ValueError, TypeError):
                                # Skip non-numeric values
                                pass

                # Store values
                advance_ids.append(advance_id)
                merchant_names.append(merchant_name)
                net_amounts.append(amount)
                data_rows += 1

                # Log occasional progress
                if data_rows % 50 == 0:
                    self.logger.info(f"Processed {data_rows} data rows so far")

            self.logger.info(
                f"Found {data_rows} valid data rows out of {row_count} rows checked"
            )

            # Create DataFrame from processed data
            processed_df = pd.DataFrame(
                {
                    "Advance ID": advance_ids,
                    "Merchant Name": merchant_names,
                    "Sum of Syn Net Amount": net_amounts,
                }
            )

            # Add standardization columns
            processed_df["Sum of Syn Gross Amount"] = processed_df[
                "Sum of Syn Net Amount"
            ]
            processed_df["Total Servicing Fee"] = 0.0

            # Log amount statistics
            self.logger.info(
                f"Amount statistics: min={processed_df['Sum of Syn Net Amount'].min()}, "
                + f"max={processed_df['Sum of Syn Net Amount'].max()}, "
                + f"mean={processed_df['Sum of Syn Net Amount'].mean()}, "
                + f"sum={processed_df['Sum of Syn Net Amount'].sum()}"
            )

            # Check for all zeros
            zero_count = (processed_df["Sum of Syn Net Amount"] == 0).sum()
            if zero_count == len(processed_df):
                self.logger.warning("All rows still have zero amounts")

                # Look for non-zero amounts in specific column range
                # Typically daily payments are in columns between AF and AP
                payment_cols = []
                for col_idx in range(32, 42):  # AF to AP (approximate range)
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    payment_cols.append(col_letter)

                self.logger.info(f"Trying specific payment columns: {payment_cols}")

                # Try recalculating from these specific payment columns
                new_amounts = []
                for row_idx, _ in enumerate(advance_ids, start=start_row):
                    row_sum = 0.0
                    for col_letter in payment_cols:
                        cell = worksheet_calculated[f"{col_letter}{row_idx}"]
                        if cell.value is not None:
                            try:
                                row_sum += float(cell.value)
                            except (ValueError, TypeError):
                                # Skip non-numeric values
                                pass
                    new_amounts.append(row_sum)

                # Update amounts if we found any non-zero values
                if any(amt != 0 for amt in new_amounts):
                    self.logger.info(
                        "Found non-zero amounts in specific payment columns"
                    )
                    processed_df["Sum of Syn Net Amount"] = new_amounts
                    processed_df["Sum of Syn Gross Amount"] = new_amounts

                    # Log new statistics
                    self.logger.info(
                        f"New statistics: min={min(new_amounts)}, "
                        + f"max={max(new_amounts)}, "
                        + f"mean={sum(new_amounts)/len(new_amounts) if new_amounts else 0}, "
                        + f"sum={sum(new_amounts)}"
                    )

            # Log sample of final data
            self.logger.info(
                f"Sample of final processed data: \n{processed_df.head().to_string()}"
            )

            return processed_df

        except Exception as e:
            self.logger.error(f"Error processing with formula handling: {str(e)}")
            raise

    def process_data(self, portfolio: str) -> pd.DataFrame:
        """Process BIG format data for the given portfolio"""
        try:
            # Try formula handling approach first
            return self.process_with_formula_handling(portfolio)

        except Exception as e:
            self.logger.error(f"Formula handling approach failed: {str(e)}")

            # Fall back to pandas-based approach
            self.logger.info("Falling back to pandas-based processing")

            # Get the correct sheet name for this portfolio
            sheet_name = self.get_portfolio_sheet_name(portfolio)
            if not sheet_name:
                raise ValueError(f"Unable to find sheet for portfolio: {portfolio}")

            # Read the Excel sheet with data_only=True to evaluate formulas
            self.logger.info(f"Reading sheet with pandas: {sheet_name}")
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)

            # Process data
            processed_df = pd.DataFrame(
                {
                    "Advance ID": df.iloc[:, 0].apply(
                        self.clean_advance_id
                    ),  # Column A (Funding Id)
                    "Merchant Name": df.iloc[:, 2]
                    .astype(str)
                    .str.strip(),  # Column C (Business Name)
                }
            )

            # Try columns in the range of AJ to AP (typical daily payment columns)
            payment_cols = []
            for col_idx in range(35, 42):  # AJ to AP (approximate range)
                if col_idx < df.shape[1]:
                    payment_cols.append(col_idx)

            if payment_cols:
                self.logger.info(f"Using payment columns indices: {payment_cols}")
                processed_df["Sum of Syn Net Amount"] = df.iloc[:, payment_cols].sum(
                    axis=1
                )
            else:
                # Last resort - use column AI
                if df.shape[1] > 34:
                    processed_df["Sum of Syn Net Amount"] = df.iloc[:, 34]  # Column AI
                else:
                    processed_df["Sum of Syn Net Amount"] = 0.0

            # Add standardization columns
            processed_df["Sum of Syn Gross Amount"] = processed_df[
                "Sum of Syn Net Amount"
            ]
            processed_df["Total Servicing Fee"] = 0.0

            # Clean up invalid rows
            processed_df = processed_df[processed_df["Advance ID"].notna()]

            return processed_df

    def process(self) -> Tuple[pd.DataFrame, float, float, float, Optional[str]]:
        """Main processing method required by BaseParser interface"""
        try:
            self.logger.info(f"Starting to process BIG file: {self.file_path}")

            # Validate format
            is_valid, error_msg = self.validate_format()
            if not is_valid:
                self.logger.error(f"Format validation failed: {error_msg}")
                return None, 0, 0, 0, error_msg

            # Detect portfolio from file
            portfolio = self.detect_portfolio()
            if not portfolio:
                self.logger.error("Could not determine portfolio from file")
                return None, 0, 0, 0, "Could not determine portfolio from file"

            # Process data for detected portfolio
            processed_df = self.process_data(portfolio)
            if processed_df is None:
                self.logger.error("Failed to process data")
                return None, 0, 0, 0, "Failed to process data"

            if len(processed_df) == 0:
                self.logger.error("No valid data found after processing")
                return None, 0, 0, 0, "No valid data found to process"

            # Calculate totals
            total_net = processed_df["Sum of Syn Net Amount"].sum()
            total_gross = total_net  # Same as net since fees are not separated
            total_fee = 0.0  # Zero since fees are not separated

            # Create pivot table
            pivot = self.create_pivot_table(
                df=processed_df,
                gross_col="Sum of Syn Gross Amount",
                net_col="Sum of Syn Net Amount",
                fee_col="Total Servicing Fee",
                index=["Advance ID", "Merchant Name"],
            )

            self.logger.info(f"BIG parser completed successfully for {portfolio}")
            self.logger.info(f"Total Net: ${total_net:,.2f}")
            self.logger.info(
                f"Generated pivot table with {len(pivot) if pivot is not None else 0} rows"
            )

            return pivot, total_gross, total_net, total_fee, None

        except Exception as e:
            error_msg = f"Error processing BIG file: {str(e)}"
            self.logger.error(error_msg)
            return None, 0, 0, 0, error_msg
