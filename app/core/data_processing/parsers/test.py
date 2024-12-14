from pathlib import Path
import pandas as pd
from datetime import datetime
import json
from openpyxl.styles import Font, PatternFill, Border, Side
# from FunderSync.app.core.data_processing.parsers.acs_vesper_parser import AcsVesperParser
from bhb_parser import BHBParser
from acs_vesper_parser import AcsVesperParser
from efin_parser import EfinParser
from Excelerate3.app.core.data_processing.parsers.clear_view_parser import ClearViewParser

def process_and_export_results(file_paths: Path | list[Path], output_dir: Path = None) -> None:
    """
    Process funder file(s) and export results to various formats.
    
    Args:
        file_paths: Path to single input CSV file or list of file paths
        output_dir: Directory for output files
    """
    try:
        # Setup output directory
        if output_dir is None:
            output_dir = file_paths[0].parent if isinstance(file_paths, list) else file_paths.parent
        else:
            output_dir = Path(output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Process the file(s)
        parser = ClearViewParser(file_paths)
        
        # Print appropriate message based on number of files
        if isinstance(file_paths, list):
            file_names = ", ".join(f.name for f in file_paths)
            print(f"\nProcessing {parser.funder_name} files: {file_names}")
        else:
            print(f"\nProcessing {parser.funder_name} file: {file_paths.name}")
        
        # Validate and process
        is_valid, error = parser.validate_format()
        if not is_valid:
            print(f"❌ Validation error: {error}")
            return
            
        pivot_table, gross, net, fee, error = parser.process()
        if error:
            print(f"❌ Error processing file: {error}")
            return

        # Create output filenames
        raw_data_file = output_dir / f"raw_data_{parser.funder_name}_{timestamp}.xlsx"
        pivot_file = output_dir / f"pivot_{parser.funder_name}_{timestamp}.xlsx"
        summary_file = output_dir / f"summary_{parser.funder_name}_{timestamp}.json"

        # Create summary before exporting files
        summary = {
            "funder": parser.funder_name,
            "processed_date": timestamp,
            "input_files": [str(f) for f in file_paths] if isinstance(file_paths, list) else str(file_paths),
            "results": {
                "total_gross": float(gross),
                "total_net": float(net),
                "total_fee": float(fee),
                "total_transactions": len(parser._df),
                "unique_merchants": int(parser._df['Merchant Name'].nunique()),
                "processed_files": {
                    "raw_data": str(raw_data_file.name),
                    "pivot_table": str(pivot_file.name)
                }
            }
        }

        # 1. Export raw data with basic formatting
        with pd.ExcelWriter(raw_data_file, engine='openpyxl') as writer:
            parser._df.to_excel(writer, sheet_name='Raw Data', index=False)
            
            # Format the raw data sheet
            worksheet = writer.sheets['Raw Data']
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 2
        
        # 2. Export pivot table with formatting
        with pd.ExcelWriter(pivot_file, engine='openpyxl') as writer:
            pivot_table.to_excel(writer, sheet_name='Pivot Table')
            
            worksheet = writer.sheets['Pivot Table']
            
            # Format headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet['A1:D1']:
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border

            # Format amount columns
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.column_letter in ['C', 'D']:  # Amount columns
                        cell.number_format = '$#,##0.00'
                    cell.border = border
            
            # Adjust column widths
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = max_length + 2
        
        # 3. Export summary to JSON
        with open(summary_file, 'w') as f:
            json.dump(summary, f, indent=4)
        
        # Print results and file locations
        print("\nFiles created successfully:")
        print(f"📊 Raw Data Excel: {raw_data_file.name}")
        print(f"    - Contains original data with formatting")
        print(f"    - All columns from source CSV")
        print(f"    - Basic formatting for readability")
        
        print(f"\n📊 Pivot Table Excel: {pivot_file.name}")
        print(f"    - Contains processed pivot table")
        print(f"    - Summarized by Advance ID and Merchant Name")
        print(f"    - Formatted amounts and totals")
        
        print(f"\n📋 Summary JSON: {summary_file.name}")
        print(f"    - Processing summary and totals")
        print(f"    - File references and metadata")
        
        print(f"\n{'='*40} Summary {'='*40}")
        print(f"Total Transactions: {len(parser._df):,}")
        print(f"Total Gross: ${gross:,.2f}")
        print(f"Total Net: ${net:,.2f}")
        print(f"Total Fee: ${fee:,.2f}")
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        raise

if __name__ == "__main__":
    # file_path = Path("eFin.csv")
    file_path = [
    Path("monday_report.csv"),
    Path("tuesday_report.csv"),
    Path("wednesday_report.csv"),
    Path("thursday_report.csv"),
    Path("friday_report.csv")
]

    output_dir = Path("outputs")
    process_and_export_results(file_path, output_dir)